# =============================================================================
# exporter.py — Génération du tableau Excel du bulletin marine (Sème)
# =============================================================================

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import config

# ---------------------------------------------------------------------------
# Palette et styles
# ---------------------------------------------------------------------------

BLUE_DARK  = "1F4E79"
BLUE_MED   = "2E75B6"
BLUE_LIGHT = "BDD7EE"
WHITE      = "FFFFFF"
BLACK      = "000000"
GRAY       = "F2F2F2"
YELLOW     = "FFCC00"
RED        = "FF0000"
GREEN_LIGHT = "E2EFDA"

DAY_COLORS = {
    0: "EBF5FB",  # Monday
    1: "EBF5EB",  # Tuesday
    2: "FEF9E7",  # Wednesday
    3: "FDEDEC",  # Thursday
    4: "F4F6F7",  # Friday
    5: "FFF3E0",  # Saturday
    6: "F3E5F5",  # Sunday
}

ENG_DAYS   = {0:"Mon.",1:"Tue.",2:"Wed.",3:"Thu.",4:"Fri.",5:"Sat.",6:"Sun."}
ENG_MONTHS = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
              7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

def thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def hfont(size=9, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def dfont(size=9, bold=False, color=BLACK):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def fmt(val, dec=1):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "—"
    return f"{val:.{dec}f}" if dec > 0 else str(int(round(val)))

# ---------------------------------------------------------------------------
# Export principal
# ---------------------------------------------------------------------------

def export_excel(df: pd.DataFrame,
                 run_datetime: datetime,
                 output_path: str = None,
                 warning_text: str = None) -> str:

    # ── Nom du fichier avec date et run ────────────────────────────────────
    # Format : Marine_forecast_DDMMYYYY_HHZ.xlsx
    date_str = run_datetime.strftime("%d%m%Y")
    run_str  = f"{run_datetime.hour:02d}Z"
    if output_path is None:
        output_path = f"D:/pipeline/Marine_forecast_{date_str}_{run_str}.xlsx"
    else:
        # Si un chemin est fourni, on y injecte quand même la date et le run
        base = output_path.replace(".xlsx", "")
        output_path = f"{base}_{date_str}_{run_str}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Marine Forecast"
    ws.sheet_view.showGridLines = False

    # ── Titre ──────────────────────────────────────────────────────────────
    # Ligne 1 : titre principal
    ws.merge_cells("A1:U1")
    ws["A1"] = (f"MARINE FORECAST — {config.POINT['name'].upper()}"
                f"   |   Run : {run_datetime.strftime('%d/%m/%Y %H:%M')} UTC"
                f"   |   SWH source : {config.SWH_SOURCE.upper()}"
                f"   |   Position : {config.POINT['lat']}°N  {config.POINT['lon']}°E")
    ws["A1"].font      = hfont(size=10)
    ws["A1"].fill      = fill(BLUE_DARK)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 20

    # Ligne 2 : warning text (sans insert_rows pour éviter les erreurs de cellules fusionnées)
    ws.merge_cells("A2:U2")
    ws["A2"] = f"⚠️  {warning_text}" if warning_text else ""
    ws["A2"].font      = Font(name="Calibri", size=9, bold=True, color="1F4E79")
    ws["A2"].fill      = PatternFill("solid", fgColor="FFF3CD") if warning_text else PatternFill("solid", fgColor=WHITE)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 24 if warning_text else 0

    # ── Groupes de colonnes (ligne 2) ──────────────────────────────────────
    # Cols : A B | C D E | F G | H | I | J | K L M | N O P | Q | R S | T | U
    groups = [
        ("A3", "B3", "Date & Time"),
        ("C3", "E3", "Wind at 10m"),
        ("F3", "G3", "Wind at 100m"),
        ("H3", "H3", "MSLP\n(hPa)"),
        ("I3", "I3", "T air\n(°C)"),
        ("J3", "J3", "SST\n(°C)"),
        ("K3", "M3", "Swell 1"),
        ("N3", "P3", "Swell 2"),
        ("Q3", "Q3", "S.W\n(m)"),
        ("R3", "S3", "Currents"),
        ("T3", "T3", "Alert"),
        ("U3", "U3", "SWH\nSource"),
    ]
    for start, end, label in groups:
        if start != end:
            ws.merge_cells(f"{start}:{end}")
        c = ws[start]
        c.value = label
        c.font  = hfont(size=9)
        c.fill  = fill(BLUE_MED)
        c.alignment = center(wrap=True)
        c.border = thin()

    # ── Sous-en-têtes (ligne 3) ────────────────────────────────────────────
    # 21 colonnes : A à U
    sub = [
        "Date", "Time",                            # A B
        "Dir.", "Spd.\n(kts)", "Gust\n(kts)",     # C D E — Wind 10m
        "Dir.", "Spd.\n(kts)",                     # F G   — Wind 100m
        "MSLP", "T (°C)", "SST\n(°C)",            # H I J
        "Dir.", "Per.\n(s)", "Ht.\n(m)",           # K L M — Swell 1
        "Dir.", "Per.\n(s)", "Ht.\n(m)",           # N O P — Swell 2
        "S.W\n(m)",                                # Q     — SWH
        "Dir.", "Spd.\n(kts)",                     # R S   — Currents
        "⚠️", "Src",                               # T U
    ]
    widths = [10, 7,  5, 6, 6,  5, 6,  7, 6, 6,  5, 5, 6,  5, 5, 6,  6,  5, 6,  5, 8]

    for i,(sh,cw) in enumerate(zip(sub,widths),1):
        col = get_column_letter(i)
        c   = ws.cell(row=4,column=i,value=sh)
        c.font      = hfont(size=8)
        c.fill      = fill(BLUE_LIGHT)
        c.alignment = center(wrap=True)
        c.border    = thin()
        c.font      = Font(name="Calibri",size=8,bold=True,color=BLUE_DARK)
        ws.column_dimensions[col].width = cw

    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 26

    # ── Données ────────────────────────────────────────────────────────────
    current_day = None
    row_idx = 5

    for _, row in df.iterrows():
        dt = pd.to_datetime(row.get("valid_local"))
        if dt is None or pd.isna(dt):
            continue

        day_key  = f"{ENG_DAYS[dt.weekday()]} {dt.day:02d} {ENG_MONTHS[dt.month]}"
        hour_str = dt.strftime("%H:%M")
        day_bg   = DAY_COLORS[dt.weekday()]

        date_disp = day_key if day_key != current_day else ""
        if day_key != current_day:
            current_day = day_key

        swh      = row.get("swh_m")
        wind_spd = row.get("wind10_spd_kt")

        # Alerte
        if swh and not pd.isna(swh) and swh >= config.ALERT_SWH_DANGER:
            alerte = "🔴"
        elif swh and not pd.isna(swh) and swh >= config.ALERT_SWH_WARNING:
            alerte = "🟡"
        elif wind_spd and not pd.isna(wind_spd) and wind_spd >= config.ALERT_WIND_DANGER:
            alerte = "🔴"
        elif wind_spd and not pd.isna(wind_spd) and wind_spd >= config.ALERT_WIND_WARNING:
            alerte = "🟡"
        else:
            alerte = "✅"

        values = [
            date_disp,
            hour_str,
            # Vent 10m
            str(row.get("wind10_dir") or "—"),
            fmt(wind_spd, 0),
            fmt(row.get("wind10_gust_kt"), 0),
            # Vent 100m
            str(row.get("wind100_dir") or "—"),
            fmt(row.get("wind100_spd_kt"), 0),
            # Paramètres météo
            fmt(row.get("mslp_hpa"), 1),
            fmt(row.get("t2m_c"), 1),
            fmt(row.get("sst_c"), 1),
            # Swell 1
            str(row.get("sw1_dir") or "—"),
            fmt(row.get("sw1_period_s"), 0),
            fmt(row.get("sw1_ht_m"), 1),
            # Swell 2
            str(row.get("sw2_dir") or "—"),
            fmt(row.get("sw2_period_s"), 0),
            fmt(row.get("sw2_ht_m"), 1),
            # SWH + courants + alerte + source
            fmt(swh, 1),
            str(row.get("cur_dir") or "—"),
            fmt(row.get("cur_spd_kt"), 1),
            alerte,
            str(row.get("swh_source") or "—"),
        ]

        for col_i, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col_i, value=val)
            c.alignment = center()
            c.border    = thin()
            c.font      = dfont()
            c.fill      = fill(day_bg)

        # Couleur cellule SWH (colonne Q = 17)
        swh_cell = ws.cell(row=row_idx, column=17)
        if swh and not pd.isna(swh):
            if swh >= config.ALERT_SWH_DANGER:
                swh_cell.fill = fill(RED)
                swh_cell.font = dfont(bold=True, color=WHITE)
            elif swh >= config.ALERT_SWH_WARNING:
                swh_cell.fill = fill(YELLOW)
                swh_cell.font = dfont(bold=True)

        # Couleur cellule vitesse vent 10m (colonne D = 4)
        wind_cell = ws.cell(row=row_idx, column=4)
        if wind_spd and not pd.isna(wind_spd):
            if wind_spd >= config.ALERT_WIND_DANGER:
                wind_cell.fill = fill(RED)
                wind_cell.font = dfont(bold=True, color=WHITE)
            elif wind_spd >= config.ALERT_WIND_WARNING:
                wind_cell.fill = fill(YELLOW)
                wind_cell.font = dfont(bold=True)

        # Couleur cellule rafale vent 10m (colonne E = 5)
        gust_cell = ws.cell(row=row_idx, column=5)
        gust = row.get("wind10_gust_kt")
        if gust and not pd.isna(gust):
            if gust >= config.ALERT_WIND_DANGER + 5:
                gust_cell.fill = fill(RED)
                gust_cell.font = dfont(bold=True, color=WHITE)
            elif gust >= config.ALERT_WIND_WARNING + 5:
                gust_cell.fill = fill(YELLOW)
                gust_cell.font = dfont(bold=True)

        ws.row_dimensions[row_idx].height = 15
        row_idx += 1

    # ── Légende ────────────────────────────────────────────────────────────
    row_idx += 1
    ws.merge_cells(f"A{row_idx}:U{row_idx}")
    ws[f"A{row_idx}"] = ("✅ Normal   🟡 Warning (SWH≥1.5m or Wind≥15kts)   "
                         "🔴 Alert (SWH≥2.0m or Wind≥20kts)")
    ws[f"A{row_idx}"].font      = Font(name="Calibri", size=8, italic=True, color="444444")
    ws[f"A{row_idx}"].alignment = Alignment(horizontal="left")

    row_idx += 1
    ws.merge_cells(f"A{row_idx}:U{row_idx}")
    ws[f"A{row_idx}"] = ("Dir=Direction  |  Per=Period(s)  |  Ht=Height(m)  |  "
                         "S.W=Significant Wave Height  |  "
                         "Sources: ECMWF Open Data + Copernicus Marine")
    ws[f"A{row_idx}"].font      = Font(name="Calibri", size=8, italic=True, color="888888")
    ws[f"A{row_idx}"].alignment = Alignment(horizontal="left")

    # ── Figer en-têtes ─────────────────────────────────────────────────────
    ws.freeze_panes = "A5"

    wb.save(output_path)
    print(f"  ✅ Fichier Excel généré : {output_path}")
    return output_path
